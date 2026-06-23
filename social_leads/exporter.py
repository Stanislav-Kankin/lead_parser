from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from storage.social_lead_repository import get_project_names_for_social_leads, get_social_leads


def export_social_leads_to_xlsx(project_id: int | None = None) -> Path:
    items = get_social_leads(limit=10000, project_id=project_id)
    project_names = get_project_names_for_social_leads([item.id for item in items])

    wb = Workbook()
    ws = wb.active
    ws.title = "people_icp"
    headers = [
        "проект",
        "score",
        "статус",
        "человек",
        "роль",
        "компания TenChat",
        "ИНН",
        "ОГРН",
        "юр. название",
        "страница компании TenChat",
        "web-компания",
        "web-сайт",
        "профиль TenChat",
        "поисковый запрос",
        "почему рейтинг",
        "ICP",
        "контекст / боль",
        "CJM",
        "заход",
        "черновик сообщения",
        "комментарий",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append(
            [
                project_names.get(item.id, ""),
                int(item.lead_score or 0),
                item.status or "",
                item.person_name or "",
                item.role_title or "",
                item.company_name or "",
                item.company_inn or "",
                item.company_ogrn or "",
                item.company_legal_name or "",
                item.company_url or "",
                item.matched_web_title or "",
                item.matched_web_domain or "",
                item.profile_url or item.source_url or "",
                item.source_query or "",
                item.why_relevant or "",
                item.likely_icp or "",
                item.pain_detected or "",
                item.cjm_stage or "",
                item.outreach_angle or "",
                item.opener or "",
                item.comment or "",
            ]
        )

    _autosize(ws)
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    scope = f"project_{project_id}_" if project_id else "all_"
    file_path = export_dir / f"people_icp_{scope}{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    return file_path


def export_social_lead_inns_to_xlsx(project_id: int | None = None) -> Path:
    items = get_social_leads(limit=10000, project_id=project_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "inn"
    headers = ["Имя", "Организация", "ИНН"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    seen_inns: set[str] = set()
    for item in items:
        inn = str(item.company_inn or "").strip()
        if not inn or inn in seen_inns:
            continue
        seen_inns.add(inn)
        ws.append(
            [
                item.person_name or "",
                item.company_legal_name or item.company_name or "",
                inn,
            ]
        )

    _autosize(ws)
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    scope = f"project_{project_id}_" if project_id else "all_"
    file_path = export_dir / f"people_inn_{scope}{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    return file_path


def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 62)
