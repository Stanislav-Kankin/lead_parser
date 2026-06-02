from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from storage.lead_repository import get_project_names_for_leads, get_web_leads, list_inns


def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)


def _site_url(item) -> str:
    domain = item.domain_normalized or item.domain or ""
    if not domain:
        return ""
    return f"https://{domain}"


def export_web_leads_to_xlsx(project_id: int | None = None) -> Path:
    items = get_web_leads(limit=10000, project_id=project_id)
    project_names = get_project_names_for_leads([item.id for item in items])

    wb = Workbook()
    ws = wb.active
    ws.title = "web_icp"

    headers = [
        "проект",
        "категория",
        "название сайта",
        "сайт",
        "контакты(телефон)",
        "почта",
        "инн",
        "юр. название",
        "огрн",
        "score",
        "тип ICP",
        "приоритет",
        "статус",
        "есть каталог",
        "есть корзина",
        "оценка ecom",
        "тип сайта",
        "оценка сайта",
        "гипотеза",
        "почему подходит",
        "заход",
        "поисковый запрос",
        "источник",
        "фокус: статус",
        "фокус: регион",
        "фокус: выручка",
        "фокус: сотрудники",
        "фокус: оквэд",
        "фокус: руководитель",
        "фокус: обновлено",
        "обновлено",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append(
            [
                project_names.get(item.id, ""),
                item.search_category or item.lead_type or item.priority or "",
                item.title or item.company_name or item.domain or "",
                _site_url(item),
                item.company_phone or "",
                item.company_email or "",
                item.company_inn or "",
                item.company_legal_name or "",
                item.company_ogrn or "",
                int(item.icp_score or 0),
                item.lead_type or "",
                item.priority or "",
                item.status or "",
                "да" if item.has_catalog else "нет",
                "да" if item.has_cart else "нет",
                int(item.ecommerce_score or 0),
                item.site_type or "",
                item.site_assessment or "",
                item.hypothesis or "",
                item.evidence or item.icp_reason or "",
                item.outreach_angle or "",
                item.query or "",
                item.source_url or "",
                item.focus_status or "",
                item.focus_region or "",
                item.focus_revenue or "",
                item.focus_employees or "",
                item.focus_okved or "",
                item.focus_director or "",
                item.focus_loaded_at.strftime("%Y-%m-%d %H:%M:%S") if item.focus_loaded_at else "",
                item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if item.updated_at else "",
            ]
        )

    _autosize(ws)

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    scope = f"project_{project_id}" if project_id else "all"
    file_path = export_dir / f"web_icp_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    return file_path


def export_inns_to_txt(project_id: int | None = None) -> Path:
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    scope = f"project_{project_id}" if project_id else "all"
    file_path = export_dir / f"inn_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.txt"
    file_path.write_text("\n".join(list_inns(project_id=project_id)), encoding="utf-8")
    return file_path
