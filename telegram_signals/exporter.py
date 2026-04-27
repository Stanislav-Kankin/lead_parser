from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from telegram_signals.repository import (
    get_discussion_leads,
    get_market_intelligence,
    get_review_leads,
    get_reviewed_leads,
    get_signals,
    get_target_leads,
)


def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)


def export_signals_to_xlsx(kind: str = "actionable") -> Path:
    if kind == "discussion":
        items = get_discussion_leads()
        suffix = "discussion"
        title = "discussion_leads"
    elif kind == "review":
        items = get_review_leads()
        suffix = "review"
        title = "review_leads"
    elif kind == "ok":
        items = get_reviewed_leads("ok")
        suffix = "ok_leads"
        title = "ok_leads"
    elif kind == "not_ok":
        items = get_reviewed_leads("not_ok")
        suffix = "not_ok_leads"
        title = "not_ok_leads"
    elif kind == "raw":
        items = get_signals(limit=None)
        suffix = "raw"
        title = "raw_signals"
    elif kind == "market":
        items = get_market_intelligence()
        suffix = "market"
        title = "market_intelligence"
    elif kind == "target":
        items = get_target_leads()
        suffix = "target"
        title = "target_leads"
    else:
        items = get_signals(limit=None, lead_fit_in=["target", "review"])
        suffix = "sales_leads"
        title = "sales_leads"

    wb = Workbook()
    ws = wb.active
    ws.title = title

    headers = [
        "date",
        "segment",
        "lead_fit",
        "next_step",
        "final_lead_score",
        "conversation_score",
        "reply_depth",
        "author_type_guess",
        "conversation_type",
        "message_type",
        "contact_entity_type",
        "is_person_reachable",
        "pain_detected",
        "icp_detected",
        "why_actionable",
        "company_hint",
        "website_hint",
        "contact_hint",
        "outreach_segment",
        "outreach_stage",
        "outreach_angle",
        "chat_title",
        "author_username",
        "short_message",
        "recommended_opener",
        "chat_url",
        "conversation_key",
        "review_status",
        "reviewed_at",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append([
            str(item.message_date)[:19] if item.message_date else "",
            item.segment or "",
            item.lead_fit or "",
            item.next_step or "",
            item.final_lead_score or 0,
            item.conversation_score or 0,
            item.reply_depth or 0,
            item.author_type_guess or "",
            item.conversation_type or "",
            item.message_type or "",
            item.contact_entity_type or "",
            1 if item.is_person_reachable else 0,
            item.pain_detected or "",
            item.icp_detected or "",
            item.why_actionable or "",
            item.company_hint or "",
            item.website_hint or "",
            item.contact_hint or "",
            item.outreach_segment or "",
            item.outreach_stage or "",
            item.outreach_angle or "",
            item.chat_title or "",
            item.author_username or "",
            (item.text_excerpt or "")[:240],
            item.recommended_opener or "",
            item.chat_url or "",
            item.conversation_key or "",
            item.review_status or "",
            str(item.reviewed_at)[:19] if item.reviewed_at else "",
        ])

    _autosize(ws)

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    file_path = export_dir / f"telegram_signals_{suffix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    return file_path
