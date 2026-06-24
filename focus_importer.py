from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import or_, select

from models.lead import Lead
from storage.db import SessionLocal
from utils.domain_normalizer import normalize_domain


def import_focus_file(path: str | Path) -> dict:
    rows = _read_rows(Path(path))
    matched = 0
    matched_by_inn = 0
    matched_by_domain = 0
    matched_by_email = 0
    matched_by_phone = 0
    unmatched = 0
    skipped = 0

    with SessionLocal() as session:
        for row in rows:
            data = _map_focus_row(row)
            inn = data.get("company_inn")
            if not inn and not (data.get("focus_website") or data.get("focus_email") or data.get("focus_phone")):
                skipped += 1
                continue

            lead, match_by = _find_matching_lead(session, data)
            if not lead:
                unmatched += 1
                continue

            if data.get("company_inn") and not lead.company_inn:
                lead.company_inn = data["company_inn"]
            if data.get("company_ogrn") and not lead.company_ogrn:
                lead.company_ogrn = data["company_ogrn"]
            if data.get("company_legal_name") and not lead.company_legal_name:
                lead.company_legal_name = data["company_legal_name"]
            lead.focus_legal_name = data.get("company_legal_name") or lead.focus_legal_name
            lead.focus_status = data.get("focus_status") or lead.focus_status
            lead.focus_region = data.get("focus_region") or lead.focus_region
            lead.focus_address = data.get("focus_address") or lead.focus_address
            lead.focus_revenue = data.get("focus_revenue") or lead.focus_revenue
            lead.focus_balance = data.get("focus_balance") or lead.focus_balance
            lead.focus_profit = data.get("focus_profit") or lead.focus_profit
            lead.focus_arbitration = data.get("focus_arbitration") or lead.focus_arbitration
            lead.focus_employees = data.get("focus_employees") or lead.focus_employees
            lead.focus_okved = data.get("focus_okved") or lead.focus_okved
            lead.focus_other_okved = data.get("focus_other_okved") or lead.focus_other_okved
            lead.focus_director = data.get("focus_director") or lead.focus_director
            lead.focus_msp = data.get("focus_msp") or lead.focus_msp
            lead.focus_phone = data.get("focus_phone") or lead.focus_phone
            lead.focus_email = data.get("focus_email") or lead.focus_email
            lead.focus_website = data.get("focus_website") or lead.focus_website
            lead.focus_registration_date = data.get("focus_registration_date") or lead.focus_registration_date
            lead.focus_loaded_at = datetime.utcnow()
            lead.updated_at = datetime.utcnow()
            matched += 1
            if match_by == "inn":
                matched_by_inn += 1
            elif match_by == "domain":
                matched_by_domain += 1
            elif match_by == "email":
                matched_by_email += 1
            elif match_by == "phone":
                matched_by_phone += 1
        session.commit()

    return {
        "rows": len(rows),
        "matched": matched,
        "matched_by_inn": matched_by_inn,
        "matched_by_domain": matched_by_domain,
        "matched_by_email": matched_by_email,
        "matched_by_phone": matched_by_phone,
        "unmatched": unmatched,
        "skipped": skipped,
    }


def read_focus_rows(path: str | Path) -> list[dict]:
    return _read_rows(Path(path))


def map_focus_row(row: dict) -> dict:
    return _map_focus_row(row)


def _find_matching_lead(session, data: dict) -> tuple[Lead | None, str | None]:
    inn = data.get("company_inn")
    if inn:
        lead = session.execute(select(Lead).where(Lead.company_inn == inn).limit(1)).scalars().first()
        if lead:
            return lead, "inn"

    domain = normalize_domain(data.get("focus_website") or "")
    if domain:
        lead = session.execute(
            select(Lead).where(
                or_(
                    Lead.domain_normalized == domain,
                    Lead.root_domain == domain,
                    Lead.domain == domain,
                    Lead.domain.ilike(f"%{domain}%"),
                )
            ).limit(1)
        ).scalars().first()
        if lead:
            return lead, "domain"

    for email in _extract_emails(data.get("focus_email")):
        lead = session.execute(select(Lead).where(Lead.company_email.ilike(f"%{email}%")).limit(1)).scalars().first()
        if lead:
            return lead, "email"

    for phone in _extract_phones(data.get("focus_phone")):
        lead = session.execute(select(Lead).where(Lead.company_phone.ilike(f"%{phone[-7:]}%")).limit(1)).scalars().first()
        if lead:
            return lead, "phone"

    return None, None


def _read_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError("Поддерживаются только .xlsx, .xlsm и .csv")


def _read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(value) for value in rows[0]]
        result = []
        for raw in rows[1:]:
            item = {headers[index]: raw[index] for index in range(min(len(headers), len(raw))) if headers[index]}
            if any(value not in (None, "") for value in item.values()):
                result.append(item)
        return result
    finally:
        wb.close()


def _read_csv(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8-sig")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [{_normalize_header(key): value for key, value in row.items()} for row in reader]


def _map_focus_row(row: dict) -> dict:
    return {
        "company_inn": _digits(_pick(row, "инн", "inn"), {10, 12}),
        "company_ogrn": _digits(_pick(row, "огрн", "ogrn"), {13, 15}),
        "company_legal_name": _pick(row, "наименование", "название", "компания", "организация", "юрназвание", "company", "name"),
        "focus_status": _pick(row, "статус", "состояние", "status"),
        "focus_region": _pick(row, "регионрегистрации", "регион", "субъект", "region"),
        "focus_address": _pick(row, "адрес", "address"),
        "focus_revenue": _format_money(_pick(row, "выручка", "доход", "revenue")),
        "focus_balance": _format_money(_pick(row, "баланс", "balance")),
        "focus_profit": _format_money(_pick(row, "чистаяприбыльубыток", "прибыль", "убыток", "profit")),
        "focus_arbitration": _format_money(_pick(row, "арбитражответчик", "арбитраж", "arbitration")),
        "focus_employees": _format_people(_pick(row, "количествосотрудников", "сотрудников", "сотрудники", "численность", "персонал", "employees", "staff")),
        "focus_okved": _pick(row, "основнойвиддеятельности", "основнойоквэд", "оквэд", "виддеятельности", "okved"),
        "focus_other_okved": _pick(row, "другиевидыдеятельности", "прочиеоквэд", "дополнительныеоквэд", "otherokved"),
        "focus_director": _pick(row, "фиоруководителя", "руководителя", "руководитель", "директор", "генеральныйдиректор", "director", "ceo"),
        "focus_msp": _pick(row, "реестрмсп", "мсп", "sme"),
        "focus_phone": _pick(row, "номертелефона", "телефон", "phone"),
        "focus_email": _pick(row, "электроннаяпочта", "почта", "email"),
        "focus_website": _pick(row, "ссылканасайт", "сайт", "website", "url"),
        "focus_registration_date": _pick(row, "датарегистрации", "registrationdate"),
    }


def _pick(row: dict, *needles: str) -> str | None:
    for key, value in row.items():
        normalized_key = _normalize_header(key)
        if any(needle in normalized_key for needle in needles):
            clean = str(value or "").strip()
            if clean:
                return clean
    return None


def _digits(value: str | None, allowed_lengths: set[int]) -> str | None:
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) in allowed_lengths:
        return digits
    return None


def _format_people(value: str | None) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    digits = re.sub(r"\D+", "", raw)
    if digits:
        return f"{int(digits):,}".replace(",", " ") + " чел."
    return raw


def _format_money(value: str | None) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    normalized_raw = raw.lower()
    if "информация отсутствует" in normalized_raw or normalized_raw in {"none", "nan"}:
        return None
    digits = re.sub(r"[^\d\-.,]+", "", raw).replace(",", ".")
    if not digits or digits in {"-", "."}:
        return raw
    try:
        amount = float(digits)
    except ValueError:
        return raw
    sign = "-" if amount < 0 else ""
    amount = abs(amount)
    if amount >= 1_000_000_000:
        value_text = _format_decimal(amount / 1_000_000_000)
        return f"{sign}{value_text} млрд ₽"
    if amount >= 1_000_000:
        value_text = _format_decimal(amount / 1_000_000)
        return f"{sign}{value_text} млн ₽"
    if amount >= 1_000:
        value_text = _format_decimal(amount / 1_000)
        return f"{sign}{value_text} тыс. ₽"
    return f"{sign}{int(amount):,}".replace(",", " ") + " ₽"


def _format_decimal(value: float) -> str:
    text = f"{value:.1f}" if value < 100 else f"{value:.0f}"
    return text.replace(".", ",")


def _extract_emails(value: str | None) -> list[str]:
    raw = str(value or "").lower()
    return list(dict.fromkeys(re.findall(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}", raw)))


def _extract_phones(value: str | None) -> list[str]:
    raw = str(value or "")
    result = []
    for candidate in re.findall(r"(?:\+?7|8)?[\s\-()]*\d[\d\s\-()]{6,}", raw):
        digits = re.sub(r"\D+", "", candidate)
        if len(digits) >= 10:
            result.append(digits[-10:])
    return list(dict.fromkeys(result))


def _normalize_header(value) -> str:
    return re.sub(r"[^a-zа-я0-9]+", "", str(value or "").lower())
